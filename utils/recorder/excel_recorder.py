# -*- coding: utf-8 -*-
# @Time    : 2021/1/3
# @Author  : Lart Pang
# @GitHub  : https://github.com/lartpang
import contextlib
import os
import re

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# Thanks:
# - Python_Openpyxl: https://www.cnblogs.com/programmer-tlh/p/10461353.html
# - Python之re模块: https://www.cnblogs.com/shenjianping/p/11647473.html


def create_xlsx(xlsx_path: str):
    if not os.path.exists(xlsx_path):
        print("We have created a new excel file!!!")
        Workbook().save(xlsx_path)
    else:
        print("Excel file has existed!")


@contextlib.contextmanager
def open_excel(xlsx_path: str, sheet_name: str):
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name, index=0)
    sheet = wb[sheet_name]

    yield sheet

    wb.save(xlsx_path)


def append_row(sheet: Worksheet, row_data):
    assert isinstance(row_data, (tuple, list))
    sheet.append(row_data)


def insert_row(sheet: Worksheet, row_data, row_id, min_col=1, interval=0):
    """
    将数据插入工作表中的一行

    Args:
        sheet: 工作表对象
        row_data: 要插入的数据，tuple或者list
        row_id: 要插入区域的行的序号（从1开始）
        min_col: 要插入区域的起始列的序号（从1开始）
        interval: row_data中各个数据之间要间隔多少个空的cell
    """
    assert isinstance(row_id, int) and isinstance(min_col, int) and row_id > 0 and min_col > 0
    assert isinstance(row_data, (tuple, list)), row_data

    num_elements = len(row_data)
    row_data = iter(row_data)
    for row in sheet.iter_rows(
            min_row=row_id, max_row=row_id, min_col=min_col, max_col=min_col + (interval + 1) * (num_elements - 1)
    ):
        for i, cell in enumerate(row):
            if i % (interval + 1) == 0:
                sheet.cell(row=row_id, column=cell.column, value=next(row_data))


def insert_cell(sheet: Worksheet, row_id, col_id, value):
    assert isinstance(row_id, int) and isinstance(col_id, int) and row_id > 0 and col_id > 0

    sheet.cell(row=row_id, column=col_id, value=value)


def merge_region(sheet: Worksheet, min_row, max_row, min_col, max_col):
    assert max_row >= min_row > 0 and max_col >= min_col > 0

    merged_region = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    sheet.merge_cells(merged_region)


def get_col_id_with_row_id(sheet: Worksheet, col_name: str, row_id):
    """
    从指定行中寻找特定的列名，并返回对应的列序号
    """
    assert isinstance(row_id, int) and row_id > 0

    for cell in sheet[row_id]:
        if cell.value == col_name:
            return cell.column
    raise ValueError(f"In row {row_id}, there is not the column {col_name}!")


def get_row_id_with_col_name(sheet: Worksheet, row_name: str, col_name: str):
    """
    从指定列名字的一列中寻找指定行，返回对应的row_id, col_id, is_new_row
    """
    is_new_row = True
    col_id = get_col_id_with_row_id(sheet=sheet, col_name=col_name, row_id=1)

    row_id = 0
    for cell in sheet[get_column_letter(col_id)]:
        row_id = cell.row
        if cell.value == row_name:
            return (row_id, col_id), not is_new_row
    return (row_id + 1, col_id), is_new_row


def get_row_id_with_col_id(sheet: Worksheet, row_name: str, col_id: int):
    """
    从指定序号的一列中寻找指定行
    """
    assert isinstance(col_id, int) and col_id > 0

    is_new_row = True
    row_id = 0
    for cell in sheet[get_column_letter(col_id)]:
        row_id = cell.row
        if cell.value == row_name:
            return row_id, not is_new_row
    return row_id + 1, is_new_row


def format_string_with_config(string: str, repalce_config: dict = None):
    assert repalce_config is not None

    if repalce_config.get("lower"):
        string = string.lower()
    elif repalce_config.get("upper"):
        string = string.upper()
    elif repalce_config.get("title"):
        string = string.title()

    if sub_rule := repalce_config.get("replace"):
        string = re.sub(pattern=sub_rule[0], repl=sub_rule[1], string=string)
    return string


class MetricExcelRecorder(object):
    def __init__(
            self,
            xlsx_path: str,
            sheet_name="results",
            row_header="methods",
            repalce_config=None,
            dataset_names=None,
            metric_names=None,
    ):
        """
        Args:
            xlsx_path: 保存工作表的xlsx文件地址
            sheet_name: 存放数据的工作表名字
            row_header: 最左上角的数据，在这个类中，指代存放于合并后的A1:A2区域的文本
            repalce_config: 用来格式化数据集名字和指标名字的设定，这里借助re.sub函数进行处理，
                默认设置：`dict(lower=True, replace=(r"[_-]", ""))`
            dataset_names: 数据集合名字列表
            metric_names: 指标名字列表
        """
        create_xlsx(xlsx_path=xlsx_path)

        if repalce_config is None:
            repalce_config = dict(lower=True, replace=(r"[_-]", ""))
        if dataset_names is None:
            dataset_names = ["pascals", "ecssd", "hkuis", "dutste", "dutomron"]
        if metric_names is None:
            metric_names = ["smeasure", "wfmeasure", "mae", "adpfm", "meanfm", "maxfm", "adpem", "meanem", "maxem"]

        self.xlsx_path = xlsx_path
        self.sheet_name = sheet_name
        self.repalce_config = repalce_config

        self.row_header = format_string_with_config(row_header, self.repalce_config)

        self.dataset_names = [format_string_with_config(s, self.repalce_config) for s in dataset_names]
        self.metric_names = [format_string_with_config(s, self.repalce_config) for s in metric_names]
        self.num_datasets = len(self.dataset_names)
        self.num_metrics = len(self.metric_names)

        self._initial_table()

    def _initial_table(self):
        """
        |-------|-------------|---------------|-----------------|---------------|-----------------|-------------------|
        |methods|dataset_name1|dataset_length1|...|dataset_name1|dataset_length1|...|dataset_name1|dataset_length1... |
        |       |metric1      |metric2        |...|metric1      |metric2        |...|metric1      |metric2...         |
        |-------|-------------|---------------|-----------------|---------------|-----------------|-------------------|
        |...
        """
        with open_excel(xlsx_path=self.xlsx_path, sheet_name=self.sheet_name) as sheet:
            # 插入row_header
            insert_cell(sheet=sheet, row_id=1, col_id=1, value=self.row_header)
            # 合并row_header的单元格
            merge_region(sheet=sheet, min_row=1, max_row=2, min_col=1, max_col=1)
            # 插入数据集信息
            insert_row(sheet=sheet, row_data=self.dataset_names, row_id=1, min_col=2, interval=self.num_metrics - 1)
            # 插入指标信息
            for i in range(self.num_datasets):
                insert_row(sheet=sheet, row_data=self.metric_names, row_id=2, min_col=2 + i * self.num_metrics)

    def _format_row_data(self, row_data: dict) -> list:
        row_data = {format_string_with_config(k, self.repalce_config): v for k, v in row_data.items()}
        return [row_data[n] for n in self.metric_names]

    def __call__(self, row_data: dict, dataset_name: str, method_name: str):
        dataset_name = format_string_with_config(dataset_name, self.repalce_config)
        assert dataset_name in self.dataset_names, f"{dataset_name} is not contained in {self.dataset_names}"

        # 1 载入数据表更新后写入新表
        with open_excel(xlsx_path=self.xlsx_path, sheet_name=self.sheet_name) as sheet:
            # 2 搜索method_name是否存在，如果存在则直接寻找对应的行列坐标，如果不存在则直接使用新行
            dataset_col_start_id = get_col_id_with_row_id(sheet=sheet, col_name=dataset_name, row_id=1)
            (method_row_id, method_col_id), is_new_row = get_row_id_with_col_name(
                sheet=sheet, row_name=method_name, col_name="methods"
            )
            # 3 插入方法名字到对应的位置
            if is_new_row:
                sheet.cell(row=method_row_id, column=method_col_id, value=method_name)
            # 4 格式化指标数据部分为合理的格式，并插入表中
            row_data = self._format_row_data(row_data=row_data)
            insert_row(sheet=sheet, row_data=row_data, row_id=method_row_id, min_col=dataset_col_start_id)


class NewMetricExcelRecorder(object):
    def __init__(
            self,
            xlsx_path: str,
            repalce_config: dict = None,
            sheet_name: str = "results",
            row_header: str = "methods",
            dataset_names: tuple = ("pascals", "ecssd", "hkuis", "dutste", "dutomron"),
            metric_names: tuple = (
                    "smeasure", "wfmeasure", "mae", "adpfm", "meanfm", "maxfm", "adpem", "meanem", "maxem"),
            dataset_lengths: tuple = (850, 1000, 4447, 5017, 5168),
            record_average: bool = True,
    ):
        assert all([isinstance(x, int) for x in dataset_lengths])
        assert len(dataset_names) == len(dataset_lengths)

        create_xlsx(xlsx_path=xlsx_path)
        self.xlsx_path = xlsx_path

        if repalce_config is None:
            self.repalce_config = dict(lower=True, replace=(r"[_-]", ""))
        else:
            self.repalce_config = repalce_config

        self.row_header = format_string_with_config(row_header, self.repalce_config)
        self.dataset_names = [format_string_with_config(s, self.repalce_config) for s in dataset_names]
        self.metric_names = [format_string_with_config(s, self.repalce_config) for s in metric_names]
        self.dataset_lengths = [float(s) for s in self.dataset_lengths]
        self.record_average = record_average

        self.num_datasets = len(self.dataset_names)
        self.num_metrics = len(self.metric_names)

        self.sheet_name = sheet_name
        self._initial_table()

    def _initial_table(self):
        """
        |-------|-------------|---------------|-----------------|---------------|-----------------|-------------------|
        |methods|dataset_name1|dataset_length1|...|dataset_name1|dataset_length1|...|dataset_name1|dataset_length1... |
        |       |metric1      |metric2        |...|metric1      |metric2        |...|metric1      |metric2...         |
        |-------|-------------|---------------|-----------------|---------------|-----------------|-------------------|
        |...
        """
        with open_excel(xlsx_path=self.xlsx_path, sheet_name=self.sheet_name) as sheet:
            # 插入row_headers
            insert_cell(sheet=sheet, row_id=1, col_id=1, value=self.row_header)
            # 合并row_header的单元格
            merge_region(sheet=sheet, min_row=1, max_row=3, min_col=1, max_col=1)

            if self.record_average:
                # 根据需要插入平均指标区域
                self.dataset_names.append("average")
                self.dataset_lengths.append(sum(self.dataset_lengths))
                self.num_datasets += 1

            # 在第一行插入数据集名字和数据量
            insert_row(sheet=sheet, row_data=self.dataset_names, row_id=1, min_col=2, interval=self.num_metrics - 1)
            insert_row(sheet=sheet, row_data=self.dataset_lengths, row_id=1, min_col=3, interval=self.num_metrics - 1)
            # 在第二行插入指标信息
            for i in range(len(self.dataset_names)):
                insert_row(sheet=sheet, row_data=self.metric_names, row_id=2, min_col=2 + i * self.num_metrics)

    def _format_row_data(self, row_data: dict) -> list:
        row_data = {format_string_with_config(k, self.repalce_config): v for k, v in row_data.items()}
        return [row_data[n] for n in self.metric_names]

    def __call__(self, row_data: dict, dataset_name: str, method_name: str):
        assert dataset_name in self.dataset_names, f"{dataset_name} is not contained in {self.dataset_names}"

        dataset_name = format_string_with_config(dataset_name, self.repalce_config)

        # 1 载入数据表，改写后存入新表
        with open_excel(xlsx_path=self.xlsx_path, sheet_name=self.sheet_name) as sheet:
            # 2 搜索method_name是否存在，如果存在则直接寻找对应的行列坐标，如果不存在则直接使用新行
            dataset_col_start_id = get_col_id_with_row_id(sheet=sheet, col_name=dataset_name, row_id=1)
            (method_row_id, method_col_id), is_new_row = get_row_id_with_col_name(
                sheet=sheet, row_name=method_name, col_name=self.row_header
            )
            # 3 插入方法名字到对应的位置
            if is_new_row:
                insert_cell(sheet=sheet, row_id=method_row_id, col_id=method_col_id, value=method_name)
            # 4 格式化指标数据部分为合理的格式，并插入表中
            row_data = self._format_row_data(row_data=row_data)
            insert_row(sheet=sheet, row_data=row_data, row_id=method_row_id, min_col=dataset_col_start_id)
